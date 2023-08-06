
import optparse
import datetime
import os
import sys
import signal
import subprocess
from pprint import pprint
import socket
import time
import json
import re
import urllib.request, urllib.parse, urllib.error
import copy
import yaml
from termcolor import colored

from openpyxl.workbook import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import NamedStyle, PatternFill, Font
from openpyxl.comments import Comment

def signal_handler(signal, frame):
    sys.exit(0)


def eprint(*args, **kwargs):
    print(*args, file=sys.stderr, **kwargs)


def run_command(command):
    proc = subprocess.Popen(command, stdout=subprocess.PIPE, shell=True)
    (out, err) = proc.communicate()
    proc.wait()
    rc = proc.poll()
    return (rc, out, err)


def command_output(command):
    (rc, out, err) = run_command(command)
    return out.rstrip('\n')


def main():
    signal.signal(signal.SIGINT, signal_handler)
    #parser = argparse.ArgumentParser(description='get-sites')
    parser = optparse.OptionParser(usage='get-sites')

    parser.add_option('--format',
                      action='store', dest='output_format', default='console',
                      help='Output format (console, json, xlsx, line')
    parser.add_option('-q', '--quiet',
                      action='store_true', dest='quiet',
                      help='Don\'t output progress')
    parser.add_option('--color',
                      action='store_true', dest='color',
                      help='Colorize output (do not use with grep)')

    parser.add_option('--root-path-excluded',
                      action='append', dest='root_paths_excluded',
                      default=['/usr/share/nginx/html',
                               '/var/www/html', '/var/www/example.com'],
                      help='')
    parser.add_option('--excluded-file',
                      action='store', dest='excluded_file', default='.excluded',
                      help='')

    # site-info args
    parser.add_option('--cached',
                      action='store_true', dest='site_info_cached',
                      help='Only attach site-info, don\'t generate')
    parser.add_option('--cache-time',
                      type=int, action='store', dest='cache_time', default=300,
                      help='Max cache age when call site-info, in seconds')
    parser.add_option('--delay',
                      type=int, action='store', dest='delay', default=0,
                      help='Delay between site-info calls, seconds')
    parser.add_option('--lock-file-path',
                      action='store', dest='lock_file_path', default='/tmp/sites-info.lock',
                      help='Path to lockfile')
    parser.add_option('--lock-file-max-age',
                      type=int, action='store', dest='lock_file_age', default=86400,
                      help='Max age of lockfile (since last modification), seconds')
    parser.add_option('--results-dir',
                      action='store', dest='results_dir', default='/var/log/site-info',
                      help='Path to directory with site-info results')
    parser.add_option('--force',
                      action='store_true', dest='force', default=False,
                      help='Ignore lockfile')

    parser.add_option('--group', '--groups',
                      action='append', dest='groups', default=['main'],
                      help='Groups of site-info tests')
    parser.add_option('--limit',
                      type=int, action='store', dest='limit', default=0,
                      help='Limit of sites')

    # export args
    parser.add_option('--sites-json',
                      action='store', dest='sites_json',
                      help='Get sites results from json')

    parser.add_option('--xlsx-path',
                      action='store', dest='xlsx_path',
                      help='xlsx destination path')

    #args = parser.parse_args()
    (args, options) = parser.parse_args()

    sites = Sites(args)

    if args.sites_json:
        with open(args.sites_json) as f:
            sites.sites = json.load(f)
    else:
        if not args.quiet:
            eprint('get-sites...')
        sites.load()

        #if args.site_info_generate:
        #sites.site_info_generate()

        if not args.site_info_cached:
            sites.site_info_generate(args.cache_time)
        sites.site_info_attach()

    sites.output()


class Sites:
    def __init__(self, args):
        self.args = args
        self.sites = []

    def get_hash(self, site):
        return site['root_path'].replace('/', '_') + '__' + '-'.join(self.args.groups)

    def site_info_attach(self):
        for k, site in enumerate(self.sites):
            json_path = self.args.results_dir + \
                '/' + self.get_hash(site) + '.json'
            if os.path.exists(json_path):
                with open(json_path) as f:
                    if not self.args.quiet:
                        eprint('open %s' % json_path)
                    self.sites[k]['site_info'] = json.load(f)

    def is_cached(self, file_path, cache_time = 300):
        if not os.path.exists(file_path):
            return False
        cached_age = time.time() - os.path.getmtime(file_path)
        if not self.args.quiet:
            eprint('%s age: %d' % (file_path, cached_age))
        return cached_age < cache_time

    def site_info_generate(self, cache_time = 300):
        commands = []
        if not os.path.exists(self.args.results_dir):
            os.mkdir(self.args.results_dir)
            os.chmod(self.args.results_dir, 0o700)

        for site in self.sites:
            json_path = '%s/%s.json' % (self.args.results_dir, self.get_hash(site))
            site_info_args = [site['root_path'], '--format', 'json']
            for group in self.args.groups:
                site_info_args.extend(['--group', group])

            if self.is_cached(json_path, cache_time):
                continue

            commands.append({
                'command': 'site-info',
                'args': site_info_args,
                'stdout': '',
                'stderr': '',
                'out_path': json_path
            });

        runner = CommandRunner(commands, args=self.args, delay=self.args.delay, lock_file_path=self.args.lock_file_path,
                               lock_file_age=self.args.lock_file_age, force=self.args.force)
        runner.run()

    def load(self):
        sites = []
        nginx_sites_dir = '/etc/nginx/sites-enabled'
        site_files = os.listdir(nginx_sites_dir)

        for site_file in site_files:
            with open(nginx_sites_dir + '/' + site_file) as f:
                content = f.read()
                for m in re.finditer(r'\s*root "?(.*?)"?;', content):
                    # TODO: commented root detect
                    if not m or not m.group(1):
                        continue
                    root_path = m.group(1)
                    if root_path in self.args.root_paths_excluded:
                        continue
                    if os.path.exists(root_path + '/' + self.args.excluded_file):
                        continue

                    site = {
                        'root_path': root_path,
                    }
                    sites.append(site)
        if self.args.limit > 0 and len(sites) > self.args.limit:
            sites = sites[:self.args.limit]
        self.sites = sites

    def site_info_dict(self, infos):
        d = {}
        for info in infos:
            d[info['name']] = info['result']
        return d

    def output(self):
        if self.args.output_format == 'console':
            for site in self.sites:
                print('')
                values = []
                for t in site['site_info']:
                    value = t['result'].encode('utf-8') if isinstance(t['result'], str) else str(t['result'])
                    value = t['result'].encode('utf-8') if isinstance(t['result'], bytes) else t['result']
                    value = str(t['result']) if isinstance(t['result'], bool) else t['result']
                    values.append(str(value))
                print('\t'.join(values))
        elif self.args.output_format == 'line':
            for site in self.sites:
                s = self.site_info_dict(site['site_info'])
                s['domain'] = s['domain'].encode('idna')
                values = []
                print('')
                out = colored(s['domain'] + ':', 'white') if self.args.color else s['domain'] + ':'
                print(out)

                for t in site['site_info']:
                    value = t['result'].encode('idna') if isinstance(t['result'], str) else str(t['result'])
                    print('%s: %s: %s' % (
                        s['domain'], # colored(site['root_path']),
                        t['name'],
                        colored(value, 'white') if self.args.color else value
                    ))
        elif self.args.output_format == 'json':
            print(json.dumps(self.sites))
        elif self.args.output_format == 'xlsx':
            table = XLSTable(self.args, self.sites)
            table.to_excel(self.args.xlsx_path)


class CommandRunner:
    def __init__(self, commands, args, delay=0, lock_file_path=None, lock_file_age=None, force=False):
        self.commands = commands
        self.args = args
        self.delay = delay
        self.lock_file = None
        self.lock_file_path = lock_file_path
        self.lock_file_age = lock_file_age
        self.force = force

    def lock_check(self):
        if not os.path.exists(self.lock_file_path):
            return True
        mtime = os.path.getmtime(self.lock_file_path)
        if time.time() - mtime < self.lock_file_age:
            return False

    def lock_write(self, str):
        log_str = '%s\t%s\n' % (time.strftime('%Y-%m-%d %H:%M:%S'), str)

        with open(self.lock_file_path, 'a') as f:
            f.write(log_str)
        if not self.args.quiet:
            eprint(log_str)

    def run(self):
        # if not self.lock_check():
        #     eprint('Lockfile %s exists and was modified at last %d seconds' %
        #            (self.lock_file_path, self.lock_file_age))
        #     if self.force:
        #         eprint('But --force flag defined, ignoring lockfile')
        #     else:
        #         sys.exit(1)

        self.lock_write('sites-info started, total commands: %d' %
                        len(self.commands))
        os.chmod(self.lock_file_path, 0o600)

        for k, command in enumerate(self.commands):
            cmd = '%s %s' % (command['command'], ' '.join(command['args']))
            self.lock_write('run command (%d/%d): %s' %
                            (k + 1, len(self.commands), cmd))

            (out, err) = self.run_command(cmd)
            self.commands[k]['stdout'] = out.decode('utf-8') if out else '';
            self.commands[k]['stderr'] = err.decode('utf-8') if err else '';
            if command['out_path']:
                with open(command['out_path'], 'w') as f:
                    f.write(out.decode('utf-8'))

            time.sleep(self.delay)

    def run_command(self, command):
        proc = subprocess.Popen(command, stdout=subprocess.PIPE, shell=True)
        (out, err) = proc.communicate()
        return (out, err)

"""
Build table from site-info results
"""
class XLSTable():
    def __init__(self, args, sites):
        self.groups = args.groups
        self.sites = sites

    def get_all_tests(self):
        tests_config = {}
        paths = [
            '/etc/site-info.yml',
            os.path.expanduser('~') + '/site-info.yml'
        ]

        for path in paths:
            if os.path.exists(path):
                with open(path, 'r') as f:
                    tests_config = yaml.load(f, Loader=yaml.FullLoader)

        if not tests_config:
            print("No site-info.yml found, or empty, aborting.")
            sys.exit(1)

        for t in tests_config['tests']:
            # config normalize
            if 'groups' in t:
                if not isinstance(t['groups'], (list)):
                    t['groups'] = [t['groups']]
            else:
                t['groups'] = []

        return tests_config

    def get_tests(self):
        tests_config = self.get_all_tests()
        groups = self.groups
        filtered = []

        if 'all' in groups:
            return tests_config['tests']

        for t in tests_config['tests']:
            # filter by groups intersection
            intersect = list(set(groups) & set(t['groups']))
            if intersect:
                filtered.append(t)
        return filtered


    """
    Each cell must have fields name, result
    May have field valid
    """
    def _excel_sheet_fill(self, sheet, rows, styles={}, row_offset=1):
        # fill all cells
        for row_idx in range(0, len(rows)):
            row = rows[row_idx]

            for col_idx in range(0, len(row)):
                col = row[col_idx]
                col_letter = get_column_letter(col_idx + 1)
                cell_row_idx = row_idx + 1 + row_offset
                cell = sheet.cell(cell_row_idx, col_idx + 1)
                cell.value = '%s' % (col['result'])

                if 'comment' in col and col['comment']:
                    cell.comment = Comment(col['comment'], '')

                # validation styles
                if 'valid' in col and col['valid'] in styles:
                    cell.style = styles[col['valid']]

                # set numeric types for some test results
                if isinstance(col['result'], (int, float, complex)):
                    #cell.value = str(col['result'])
                    cell.data_type = cell.TYPE_NUMERIC
                    #cell.number_format = '0.00E+00'

                # set numeric types for some test results
                elif isinstance(col['result'], str) and col['result'].isdigit():
                    #cell.value = float(col['result']) if '.' in col['result'] else int(col['result'])
                    cell.data_type = cell.TYPE_NUMERIC

        #return sheet

    """
    Autofilters, fixed columns and other beautifiers
    """

    def _excel_sheet_tune(self, sheet, style=None):
        last_result_row_idx = sheet.max_row
        last_result_col_idx = sheet.max_column

        # autofilter
        sheet_dimensions_name = 'A1:%s%s' % (get_column_letter(
            last_result_col_idx), last_result_row_idx)
        sheet.auto_filter.ref = sheet_dimensions_name

        # autosum
        sum_row_idx = sheet.max_row + 2
        sheet['A%s' % sum_row_idx].value = 'totals'
        for col_idx in range(2, last_result_col_idx):
            cell_letter = get_column_letter(col_idx)
            cell = sheet.cell(sum_row_idx, col_idx)
            column_range_name = '%s%s:%s%s' % (
                cell_letter, 2, cell_letter, last_result_row_idx)
            cell.value = '=SUM(%s)' % column_range_name

        # width
        sheet.column_dimensions['A'].width = 25
        sheet.column_dimensions['B'].width = 5
        for col_idx in range(2, last_result_col_idx):
            cell_letter = get_column_letter(col_idx)
            sheet.column_dimensions[cell_letter].width = 15

        # fixed first row and first column
        sheet.freeze_panes = sheet['B2']

        # default cell style
        # if style:
        #     for col_idx in range(last_result_col_idx):
        #         col_name = get_column_letter(col_idx + 1)
        #         col_dimensions = sheet.column_dimensions[col_name]
        #         col_dimensions.style = style

    def to_excel(self, xlsx_path):
        sites = copy.deepcopy(self.sites)
        tests_config = self.get_tests()
        book = Workbook()

        styles = {
            'default': NamedStyle(
                name='default', font=Font(name='Courier New', size=10)),
            'pass': NamedStyle(name='pass',
                            fill=PatternFill(
                                fill_type='solid', start_color='C6EFCE', end_color='C6EFCE'),
                            font=Font(color='006100')
                        ),
            'warn': NamedStyle(name='warn',
                            fill=PatternFill(
                                fill_type='solid', start_color='FFEB9C', end_color='FFEB9C'),
                            font=Font(color='9C6500')
                        ),
            'fail': NamedStyle(name='fail',
                            fill=PatternFill(
                                fill_type='solid', start_color='FFC7CE', end_color='FFC7CE'),
                            font=Font(color='9C0006')
                        )
        }
        for name,style in styles.items():
            book.add_named_style(style)

        # Sheet "drupals"
        sheet = book.active
        sheet.title = 'site-info'
        rows = []

        # header fill
        field_names = []
        for col_idx, t in enumerate(tests_config):
            col_name = get_column_letter(col_idx + 1)
            field_names.append(t['name'])
            cell = sheet.cell(1, col_idx + 1)
            cell.value = t['name']
            if 'comment' in t:
                cell.comment = Comment(t['comment'], '')

            # for test_config in self.tests_config:
            #     if test_config['name'] == col['name'] and 'comment' in test_config:
            #         #print 'comment for %s: %s' % (col['name'], test_config['comment'])
            #         cell.comment = Comment(test_config['comment'], '')

        # fill absent tests with empty values
        for site in sites:
            row = []
            site_info = site['site_info']

            # site_info
            named_results = {}
            for t in site_info:
                named_results[t['name']] = t
            for name in field_names:
                if name in named_results:
                    row.append(named_results[name])
                else:
                    row.append({'result': ''})
            rows.append(row)

        self._excel_sheet_fill(sheet, rows, styles)
        self._excel_sheet_tune(sheet, style=styles['default'])

        # test groups sheets
        # test_groups = self.get_test_groups()
        # for test_group in test_groups:
        #     sheet = book.create_sheet(title=test_group)
        #     rows = []
        #     for test_suite in test_suites:
        #         rows.append(test_suite.get_data(group=test_group))

        #     self._excel_sheet_fill(sheet, rows, styles)
        #     self._excel_sheet_tune(sheet, style=sheet_style)

        book.save(xlsx_path)


    """ not used """
    # def _list_times(self):
    #     # Sheet "times"
    #     # Need for slow tests detection
    #     sheet = book.create_sheet(title='times')
    #     rows = []
    #     for test_suite in test_suites:
    #         row_data = copy.deepcopy(test_suite.get_data())
    #         for col_data in row_data:
    #             if col_data['name'] not in ['domain']:
    #                 time = col_data['time'] if 'time' in col_data else ''
    #                 col_data['result'] = time
    #                 if time == '':
    #                     col_data['valid'] = None
    #                 else:
    #                     if time < 0.5:
    #                         col_data['valid'] = 'pass'
    #                     elif time < 2:
    #                         col_data['valid'] = 'warn'
    #                     else:
    #                         col_data['valid'] = 'fail'
    #         rows.append(row_data)

    #     self._excel_sheet_fill(sheet, rows, styles)
    #     self._excel_sheet_tune(sheet, style=sheet_style)


if __name__ == '__main__':
    main()
