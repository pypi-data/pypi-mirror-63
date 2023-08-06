#!/usr/bin/env python



import openpyxl
import os
import csv
from xlsxwriter.utility import xl_cell_to_rowcol

class  AutoSphereExcel():
    """
        This test library internally use openpyxl module of python and provides keywords to open, read, write excel files. This library only supports
        xlsx file formats.


        *Prerequisties*

        Openpyxl module of python should be installed using command "pip install openpyxl"
        AutoSphereExcel must be imported.

        Example:
            | Library        | AutoSphereExcel        |
            | Open Excel     | Filename with fullpath |

        """

    def __init__(self):
        self.wb = None
        self.sheet = None
        self.filename = None
        
    def open_excel(self, file):
        """
        Open excel file
        Arguments:
            | File             | Filename with fullpath to open and test upon        |

        Example:
        | Open Excel      |  C:\\Python27\\ExcelRobotTest\\ExcelRobotTest.xlsx  |
        """
        if os.path.exists(file):
            self.filename = file
            self.wb = openpyxl.load_workbook(self.filename)
        else:
            print('11111')
            self.wb = openpyxl.Workbook()

    def get_sheet_names(self, file):
        """
        Return sheetnames of the workbook
        Example:
        | Openexcel File       |  C:\\Python27\\ExcelRobotTest\\ExcelRobotTest.xlsx  |
        | Get sheet names      |                                                     |
        """
        self.filename = file
        return self.wb.sheetnames
    
    
    def opensheet_byname(self, sheetname):
        """
        **** Marked for depreciation ****
        """
        #self.sheet = self.wb.get_sheet_by_name(sheetname)
        self.sheet = self.wb[sheetname]

    
    def get_column_count(self, sheetname):
        """
        Return the column count of the given sheet
        Example:
        | Get Column count     |  Sheet1 |
        """
        #self.sheet = self.wb.get_sheet_by_name(sheetname)
        self.sheet = self.wb[sheetname]
        return self.sheet.max_column


    def get_row_count(self, sheetname):
        """
        Return the Row count of the given sheet
        Example:
        | Get Row count     |  Sheet1 |
        """
        #self.sheet = self.wb.get_sheet_by_name(sheetname)
        self.sheet = self.wb[sheetname]
        return self.sheet.max_row

    def read_cell_data_by_coordinates(self,sheetname, row_value, column_value):
        """
        Return the value of a cell by giving the sheetname, row value & column value
        Example:
        | Read Cell Data By Coordinates     |  SheetName | Row Number |  Column Number  |
        | Read Cell Data By Coordinates     |  Sheet1 |  1  |  1  |
        """
        #self.sheet = self.wb.get_sheet_by_name(sheetname)
        self.sheet = self.wb[sheetname]
        self.row = int(row_value)
        self.column = int(column_value)
        varcellValue =  self.sheet.cell(row=self.row, column=self.column).value
        return varcellValue

    
    def write_data_by_coordinates(self,sheetname,column_value,row_value,varValue):
        """
        Write the value to a call using its co-ordinates
        Example:
        | Write Data By Coordinates    |  SheetName  | Row Number | Column Number |  Data  |
        | Write Data By Coordinates    | Sheet1 | 1 | 1 |  TestData  |
        """
        #self.sheet = self.wb.get_sheet_by_name(sheetname)
        self.sheet = self.wb[sheetname]
        self.row = int(row_value)
        self.column = int(column_value)
        self.varValue = varValue
        self.sheet.cell(row=self.row, column=self.column).value = self.varValue
    

    def save_excel(self, file):
        """
        Save the excel file after writing the data.
        Example:
        Update existing file:

        | Openexcel File       |  C:\\Python27\\ExcelRobotTest\\ExcelRobotTest.xlsx  |
        | Save Excelfile       |  C:\\Python27\\ExcelRobotTest\\ExcelRobotTest.xlsx  |

        Save in new file:
        | Openexcel File       |  C:\\Python27\\ExcelRobotTest\\ExcelRobotTest.xlsx  |
        | Save Excelfile       |  D:\\Test\\ExcelRobotNewFile.xlsx                   |   
        """
        self.file = file
        self.wb.save(self.file)

    def add_new_sheet(self, varnewsheetname):
        """
        Add new sheet
        Arguments:
        | New sheetname        | The name of the new sheet to be added in the workbook     |

        Example:
        | Keywords             | Parameters                                       |
        | Add new sheet        | SheetName                                       |
        """
        self.newsheet = varnewsheetname
        self.wb.create_sheet(self.newsheet)

    def append_new_data(self,sheetname,data):
        """
        append_new_data
        Arguments:
        | sheetname        | the sheetname where to be appended     |
        | data             | the data to be appended

        Example:
        | Keywords         | Parameters                             |
        | Add new sheet    | SheetName                              |
        """

        self.sheet = self.wb[sheetname]
        self.sheet.append(list(data))

    def close(self):
        self.wb.close()
    def remove(self, sheetname):
        self.worksheet = sheetname
        self.wb.remove_sheet(self.worksheet)
    def get_sheet_by_name(self, sheetname):
        return self.wb.get_sheet_by_name(sheetname)              
    def print_area(self, value,sheetname):
        self.sheet = self.wb[sheetname]        
        self.sheet.print_area=value
    def paste_range(self, ref, sheetReceiving, copiedData):
        opxl=AutoSphereExcel()
        s_ref=(opxl.splitReference(ref));
        row_col=(opxl.convertToRowsColumns(s_ref[0]))
        startRow = row_col[0] + 1
        startCol = row_col[1] + 1
        row_col=(opxl.convertToRowsColumns(s_ref[1]))
        endRow = row_col[0] + 1
        endCol = row_col[1] + 1                
        self.sheet = self.wb[sheetReceiving]        
        print(copiedData)
        countRow = 0
        for i in range(int(startRow),int(endRow)+1,1):
            countCol = 0
            for j in range(int(startCol),int(endCol)+1,1):            
                self.sheet.cell(row = i, column = j).value = copiedData[countRow][countCol]
                countCol += 1
            countRow += 1
    def export_workbook_to_csv(self, workbook, csvfile):
        wb = openpyxl.load_workbook(workbook)
        sh = wb.active
        with open(csvfile, 'w', newline="") as f:  # open('test.csv', 'w', newline="") for python 3
            c = csv.writer(f)
            for r in sh.rows:
                c.writerow([cell.value for cell in r])
    def copy_range(self, ref):

        opxl=AutoSphereExcel()
        s_ref=(opxl.splitReference(ref));
        row_col=(opxl.convertToRowsColumns(s_ref[0]))
        startRow = row_col[0] + 1
        startCol = row_col[1] + 1
        row_col=(opxl.convertToRowsColumns(s_ref[1]))
        endRow = row_col[0] + 1
        endCol = row_col[1] + 1        
        rangeSelected = []
        #Loops through selected Rows
        for i in range(int(startRow),int(endRow) + 1,1):
            #Appends the row to a RowSelected list
            rowSelected = []
            for j in range(int(startCol),int(endCol)+1,1):
                rowSelected.append(self.sheet.cell(row = i, column = j).value)                
            #Adds the RowSelected List and nests inside the rangeSelected
            rangeSelected.append(rowSelected)
        return rangeSelected
          #inserting column
    def insert_column(self, col_index, offset=1):
        sh = self.sheet
        sh.insert_cols(int(col_index),offset)
    def delete_column(self, col_index, offset=1):
        sh = self.sheet
        sh.delete_cols(int(col_index),offset)
    def insert_row(self, col_index, offset=1):
        sh = self.sheet
        sh.insert_rows(int(col_index),offset)
    def delete_row(self, col_index, offset=1):
        sh = self.sheet
        sh.delete_rows(int(col_index),offset)
    def add_image(self, ref, imagepath):
        sh = self.sheet
        img = openpyxl.drawing.image.Image(imagepath)
        img.anchor = ref
        sh.add_image(img)
    def read_range(self, ref):
        opxl=AutoSphereExcel()
        s_ref=(opxl.splitReference(ref));
        row_col=(opxl.convertToRowsColumns(s_ref[0]))
        startRow = row_col[0] + 1
        startCol = row_col[1] + 1
        row_col=(opxl.convertToRowsColumns(s_ref[1]))
        endRow = row_col[0] + 1
        endCol = row_col[1] + 1        
        sheet = self.sheet #Add Sheet name
        rangeSelected = []
        #Loops through selected Rows
        for i in range(int(startRow),int(endRow) + 1,1):
            #Appends the row to a RowSelected list
            rowSelected = []
            for j in range(int(startCol),int(endCol)+1,1):
                rowSelected.append(sheet.cell(row = i, column = j).value)                
            #Adds the RowSelected List and nests inside the rangeSelected
            rangeSelected.append(rowSelected)
        return rangeSelected
    def read_column(self, startRow, startCol, endCol):
        rangeSelected = []
        #Loops through selected Rows
        for i in range(int(startCol),int(endCol) + 1,1):
            rangeSelected.append(self.sheet.cell(row = int(startRow), column = i).value)
        return rangeSelected
    def read_row(self, rowIndex):
        selected_row=self.sheet[rowIndex]
        rowSelected = []
        for i in range(int(rowIndex),(self.sheet.max_column) + 1,1):
            #Appends the row to a RowSelected list
            rowSelected.append(self.sheet.cell(row = int(rowIndex), column = i).value)   
        return rowSelected   
    def add_table(self, displayName, ref1, ref2):
        mediumStyle = openpyxl.worksheet.table.TableStyleInfo(name='TableStyleMedium2',
                                                      showRowStripes=True)
            # create a table
        refr=ref1+":"+ref2    
        table = openpyxl.worksheet.table.Table(ref=refr,
                                       displayName='FruitColors',
                                       tableStyleInfo=mediumStyle)     
        self.sheet.add_table(table)                                   

    def refresh_pivot_table(self):
            # create a table
        pivot = self.sheet._pivots[0] # any will do as they share the same cache
        pivot.cache.refreshOnLoad = True
    def filter_table(self, displayName, ref1, ref2, filterStr):
            # create a table
        ws=self.sheet
        ref=ref1+":"+ref2
        filterStr=list(filterStr)        
        ws.auto_filter.ref = ref
        ws.auto_filter.add_filter_column(0, filterStr)
        ws.auto_filter.add_sort_condition(ref)
    def append_to_csv(self, csvfile, fields):
            # create a table
        with open(r''+csvfile, 'a') as f:
            writer = csv.writer(f)
            writer.writerow(fields)
    def str2bool1(self):
        print("Hello World!")

    def str2bool(self, v):
        opxl=AutoSphereExcel()
        opxl.str2bool1()
        return v.lower() in ("yes", "true", "t", "1")

    def move_cell_range(self, ref1, ref2, rows, cols, translate):
            # create a table
        ws=self.sheet
        ref=ref1+":"+ref2
        isTranslate = (translate.lower() in ("yes", "true", "t", "1"))
        ws.move_range(ref, int(rows), int(cols))    

    def read_csv(self, csvfile, excelfile):
        wb = openpyxl.Workbook()
        ws = wb.active

        with open(csvfile) as f:
            reader = csv.reader(f, delimiter=':')
            for row in reader:
                ws.append(row)

        wb.save(excelfile)
    
    def remove_duplicate_rows(self):
        sh=self.sheet
        rangeSelected = []
        rowindex = []
    
    #thisset.clear()
        for i in range(1,int(sh.max_row) + 1,1):
            print("Processing: " + str(i))
            rowSelected = []
        
            for j in range(1,int(sh.max_column)+1,1):
                rowSelected.append(sh.cell(row = i, column = j).value)
            
            if rowSelected in rangeSelected:
                rowindex.append(i)
#               sh.delete_rows(i,1)          
            else:    
                rangeSelected.append(rowSelected)
        m = 0
        for k in rowindex:
            print(str(k) +" have duplicate entry")
            sh.delete_rows(k-m)
            m = m+1

        return rangeSelected
    def toSingleQuote(s1):
        return "'%s'" % s1                                           
    def splitReference(self, ref):
        split_ref=ref.split(":");
        return (split_ref)

    def convertToRowsColumns(self, ref):
        thistuple =xl_cell_to_rowcol(ref) 
        return thistuple

#
if __name__ == "__main__":
    opxl=AutoSphereExcel()
    opxl.open_excel('test1.xlsx')
    opxl.get_sheet_names('test1.xlsx')
    opxl.save_excel('test1.xlsx')
    print(opxl.str2bool('yes'))
    s_ref=(opxl.splitReference("A2:C5"));
    print(s_ref[0]);print(s_ref[1]);
    row_col=(opxl.convertToRowsColumns(s_ref[0]));
    print(row_col[0]);print(row_col[1]);
    row_col=(opxl.convertToRowsColumns(s_ref[1]));
    print(row_col[0]);print(row_col[1]);
    opxl.close()