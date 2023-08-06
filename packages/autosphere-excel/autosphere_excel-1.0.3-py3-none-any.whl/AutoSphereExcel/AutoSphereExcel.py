#!/usr/bin/env python



import openpyxl
import os
import csv
from xlsxwriter.utility import xl_cell_to_rowcol
from openpyxl.comments import Comment
from openpyxl import Workbook
from openpyxl.workbook.protection import WorkbookProtection
from openpyxl.worksheet.protection import SheetProtection
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell
from openpyxl.styles import NamedStyle
import os, os.path
import win32com.client

class  AutoSphereExcel():
    """
        This test library internally use openpyxl module of python and provides keywords to open, read, write excel files. This library only supports
        xlsx file formats.


        *Prerequisties*

        Openpyxl module of python should be installed using command "pip install openpyxl"
        AutoSphereExcel must be imported.

        Example:
            | Library        | AutoSphereExcel        |
            | Open Excel     | Filename with fullpath |

        """

    def __init__(self):
        self.wb = None
        self.sheet = None
        self.filename = None
        self.ws = None        
    def open_excel(self, file):
        """
        Open excel file
        Arguments:
            | File             | Filename with fullpath to open and test upon        |

        Example:
        | Open Excel      |  C:\\Python27\\ExcelRobotTest\\ExcelRobotTest.xlsx  |
        """
        if os.path.exists(file):
            self.filename = file
            self.wb = openpyxl.load_workbook(self.filename)
        else:
            self.wb = openpyxl.Workbook()

    def get_sheet_names(self):
        """
        Return sheetnames of the workbook
        Example:
        | Openexcel File       |  C:\\Python27\\ExcelRobotTest\\ExcelRobotTest.xlsx  |
        | Get sheet names      |                                                     |
        """
        # self.filename = file
        return self.wb.get_sheet_names()
    
    
    def open_sheet_by_name(self, sheetname):
        """
        **** Marked for depreciation ****
        """
        #self.sheet = self.wb.get_sheet_by_name(sheetname)
        self.sheet = self.wb[sheetname]

    
    def get_column_count(self, sheetname):
        """
        Return the column count of the given sheet
        Example:
        | Get Column count     |  Sheet1 |
        """
        #self.sheet = self.wb.get_sheet_by_name(sheetname)
        self.sheet = self.wb[sheetname]
        return self.sheet.max_column


    def get_row_count(self, sheetname):
        """
        Return the Row count of the given sheet
        Example:
        | Get Row count     |  Sheet1 |
        """
        #self.sheet = self.wb.get_sheet_by_name(sheetname)
        self.sheet = self.wb[sheetname]
        return self.sheet.max_row

    def read_cell(self,sheetname, ref):
        """
        Return the value of a cell by giving the sheetname, row value & column value
        Example:
        | Read Cell Data By Coordinates     |  SheetName | Row Number |  Column Number  |
        | Read Cell Data By Coordinates     |  Sheet1 |  1  |  1  |
        """
        opxl=AutoSphereExcel()
        row_col=(opxl.convertToRowsColumns(ref))
        row_value = row_col[0] + 1
        column_value = row_col[1] + 1

        #self.sheet = self.wb.get_sheet_by_name(sheetname)
        self.sheet = self.wb[sheetname]
        self.row = int(row_value)
        self.column = int(column_value)
        varcellValue =  self.sheet.cell(row=self.row, column=self.column).value
        return varcellValue

    
    def write_cell(self, sheetname, ref, varValue):
        """
        Write the value to a call using its co-ordinates
        Example:
        | Write Data By Coordinates    |  SheetName  | Row Number | Column Number |  Data  |
        | Write Data By Coordinates    | Sheet1 | 1 | 1 |  TestData  |
        """
        opxl=AutoSphereExcel()
        row_col=(opxl.convertToRowsColumns(ref))
        row_value = row_col[0] + 1
        column_value = row_col[1] + 1     
        #self.sheet = self.wb.get_sheet_by_name(sheetname)
        self.sheet = self.wb[sheetname]
        self.row = int(row_value)
        self.column = int(column_value)
        self.varValue = varValue
        self.sheet.cell(row=self.row, column=self.column).value = self.varValue
    

    def save_excel(self, file):
        """
        Save the excel file after writing the data.
        Example:
        Update existing file:

        | Openexcel File       |  C:\\Python27\\ExcelRobotTest\\ExcelRobotTest.xlsx  |
        | Save Excelfile       |  C:\\Python27\\ExcelRobotTest\\ExcelRobotTest.xlsx  |

        Save in new file:
        | Openexcel File       |  C:\\Python27\\ExcelRobotTest\\ExcelRobotTest.xlsx  |
        | Save Excelfile       |  D:\\Test\\ExcelRobotNewFile.xlsx                   |   
        """
        self.file = file
        self.wb.save(self.file)

    def add_new_sheet(self, varnewsheetname):
        """
        Add new sheet
        Arguments:
        | New sheetname        | The name of the new sheet to be added in the workbook     |

        Example:
        | Keywords             | Parameters                                       |
        | Add new sheet        | SheetName                                       |
        """
        self.newsheet = varnewsheetname
        self.wb.create_sheet(self.newsheet)

    def append_new_data(self,sheetname,data):
        """
        append_new_data
        Arguments:
        | sheetname        | the sheetname where to be appended     |
        | data             | the data to be appended

        Example:
        | Keywords         | Parameters                             |
        | Add new sheet    | SheetName                              |
        """

        self.sheet = self.wb[sheetname]
        self.sheet.append(list(data))

    def close(self):
        self.wb.close()
    def remove_sheet(self, sheetname):
        self.worksheet = sheetname
        self.wb.remove_sheet(self.worksheet)
    def get_sheet_by_name(self, sheetname):
        return self.wb.get_sheet_by_name(sheetname)              

#   def add_print_area(self, value):
# self.wb.active("0")
#       self.ws = openpyxl.worksheet.worksheet.Worksheet
#       self.ws.print_area(value)
#       return self.ws.print_area()
# 
    def add_comment_to_cell(self, comment, author, cellcoordinates):
        wb = self.wb
        sheet = wb.active
        # comment = sheet['A1'].comment
        comment = Comment(comment, author)
        sheet[cellcoordinates].comment = comment

    def copy_sheet(self, sheetname, newsheetname):
        wb = self.wb
        source = wb.active
        target = wb.copy_worksheet(source, newsheetname)
        return target

    def create_new_excel_file(self, workbookname, sheetname):
        wb = Workbook()
        ws = wb.active
        ws.title = sheetname
        wb.save(workbookname)

    def delete_column(self, colrange):
        wb = self.wb
        ws = wb.active
        x = colrange.split(", ")
        ws.delete_cols(int(x[0].strip()), int(x[1].strip()))

#    def create_table(self, data, headings, excelfile, coordinates, tablename):
#        wb = Workbook()
#        ws = wb.active
#        headings = headings.strip('][').split(', ')
#        data =  data.strip('][').split(', ')
        # add column headings. NB. these must be strings
#        ws.append(headings)
#        for row in data:
#            ws.append(row.strip('][').split(', '))

#        tab = Table(displayName=tablename, ref=coordinates)

        # Add a default style with striped rows and banded columns
        # style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
        #                        showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        # tab.tableStyleInfo = style

#        ws.add_table(tab)
#        wb.save(excelfile)

    def hide_column(self, col):
            wb = self.wb
            ws = wb.active
            ws.column_dimensions[col].hidden = True

    def unhide_column(self, col):
            wb = self.wb
            ws = wb.active
            ws.column_dimensions[col].hidden = False

    def hide_sheet(self):
            self.sheet.sheet_state = 'hidden'

    def unhide_sheet(self):
            self.sheet.sheet_state = 'visible'

    def set_tab_color(self, color):
            self.sheet.sheet_properties.tabColor = color
    def fill_range_color(self, ref, color):
        opxl=AutoSphereExcel()
        s_ref=(opxl.splitReference(ref));
        row_col=(opxl.convertToRowsColumns(s_ref[0]))
        startRow = row_col[0] + 1
        startCol = row_col[1] + 1
        row_col=(opxl.convertToRowsColumns(s_ref[1]))
        endRow = row_col[0] + 1
        endCol = row_col[1] + 1        
        redFill = PatternFill(start_color=color,
                          end_color=color,
                          fill_type='solid')
        countRow = 0
        print("startCol:"+startCol+", startRow:"+startRow+", endCol:"+endCol+", endRow:"+endRow+", color:"+color)
        for i in range(int(startRow), int(endRow) + 1, 1):
            print("Row:" + str(i))
            countCol = 0
            for j in range(int(startCol), int(endCol) + 1, 1):
                col = self.colnum_string(j)
                row = str(i)
                cell = col + row
                print("Cell:" + cell)
                self.sheet[cell].fill = redFill
                countCol += 1
            countRow += 1

    def set_date_format(self):
        date_style = NamedStyle(name='datetime', number_format='DD/MM/YYYY')
        self.sheet['A13'].style = date_style

    def paste_range(self, ref, sheetReceiving, copiedData):
        opxl=AutoSphereExcel()
        s_ref=(opxl.splitReference(ref));
        row_col=(opxl.convertToRowsColumns(s_ref[0]))
        startRow = row_col[0] + 1
        startCol = row_col[1] + 1
        row_col=(opxl.convertToRowsColumns(s_ref[1]))
        endRow = row_col[0] + 1
        endCol = row_col[1] + 1                
        self.sheet = self.wb[sheetReceiving]        
        print(copiedData)
        countRow = 0
        for i in range(int(startRow),int(endRow)+1,1):
            countCol = 0
            for j in range(int(startCol),int(endCol)+1,1):            
                self.sheet.cell(row = i, column = j).value = copiedData[countRow][countCol]
                countCol += 1
            countRow += 1
    def export_workbook_to_csv(self, workbook, csvfile):
        wb = openpyxl.load_workbook(workbook)
        sh = wb.active
        with open(csvfile, 'w', newline="") as f:  # open('test.csv', 'w', newline="") for python 3
            c = csv.writer(f)
            for r in sh.rows:
                c.writerow([cell.value for cell in r])
    def copy_range(self, ref):

        opxl=AutoSphereExcel()
        s_ref=(opxl.splitReference(ref));
        row_col=(opxl.convertToRowsColumns(s_ref[0]))
        startRow = row_col[0] + 1
        startCol = row_col[1] + 1
        row_col=(opxl.convertToRowsColumns(s_ref[1]))
        endRow = row_col[0] + 1
        endCol = row_col[1] + 1        
        rangeSelected = []
        #Loops through selected Rows
        for i in range(int(startRow),int(endRow) + 1,1):
            #Appends the row to a RowSelected list
            rowSelected = []
            for j in range(int(startCol),int(endCol)+1,1):
                rowSelected.append(self.sheet.cell(row = i, column = j).value)                
            #Adds the RowSelected List and nests inside the rangeSelected
            rangeSelected.append(rowSelected)
        return rangeSelected
          #inserting column
    def insert_column(self, col_index, offset=1):
        sh = self.sheet
        sh.insert_cols(int(col_index),offset)
    def delete_column(self, col_index, offset=1):
        sh = self.sheet
        sh.delete_cols(int(col_index),offset)
    def insert_row(self, row_index, offset=1):
        sh = self.sheet
        sh.insert_rows(int(row_index),offset)
    def delete_row(self, row_index, offset=1):
        sh = self.sheet
        sh.delete_rows(int(row_index),offset)
    def add_image(self, ref, imagepath):
        sh = self.sheet
        img = openpyxl.drawing.image.Image(imagepath)
        img.anchor = ref
        sh.add_image(img)
    def read_range(self, ref):
        opxl=AutoSphereExcel()
        s_ref=(opxl.splitReference(ref));
        row_col=(opxl.convertToRowsColumns(s_ref[0]))
        startRow = row_col[0] + 1
        startCol = row_col[1] + 1
        row_col=(opxl.convertToRowsColumns(s_ref[1]))
        endRow = row_col[0] + 1
        endCol = row_col[1] + 1        
        sheet = self.sheet #Add Sheet name
        rangeSelected = []
        #Loops through selected Rows
        for i in range(int(startRow),int(endRow) + 1,1):
            #Appends the row to a RowSelected list
            rowSelected = []
            for j in range(int(startCol),int(endCol)+1,1):
                rowSelected.append(sheet.cell(row = i, column = j).value)                
            #Adds the RowSelected List and nests inside the rangeSelected
            rangeSelected.append(rowSelected)
        return rangeSelected
    def read_column(self, startRow, startCol, endCol):
        rangeSelected = []
        #Loops through selected Rows
        for i in range(int(startCol),int(endCol) + 1,1):
            rangeSelected.append(self.sheet.cell(row = int(startRow), column = i).value)
        return rangeSelected
    def read_row(self, rowIndex):
        selected_row=self.sheet[rowIndex]
        rowSelected = []
        for i in range(int(rowIndex),(self.sheet.max_column) + 1,1):
            #Appends the row to a RowSelected list
            rowSelected.append(self.sheet.cell(row = int(rowIndex), column = i).value)   
        return rowSelected   
    def add_table(self, displayName, ref1, ref2):
        mediumStyle = openpyxl.worksheet.table.TableStyleInfo(name='TableStyleMedium2',
                                                      showRowStripes=True)
            # create a table
        refr=ref1+":"+ref2    
        table = openpyxl.worksheet.table.Table(ref=refr,
                                       displayName='FruitColors',
                                       tableStyleInfo=mediumStyle)     
        self.sheet.add_table(table)                                   

    def refresh_pivot_table(self):
            # create a table
        pivot = self.sheet._pivots[0] # any will do as they share the same cache
        pivot.cache.refreshOnLoad = True
    def filter_table(self, displayName, ref, filterStr):
            # create a table
        ws=self.sheet
#        ref=ref1+":"+ref2
        filterStr=list(filterStr)        
        ws.auto_filter.ref = ref
        ws.auto_filter.add_filter_column(0, filterStr)
        ws.auto_filter.add_sort_condition(ref)
    def append_to_csv(self, csvfile, fields):
            # create a table
        with open(r''+csvfile, 'a') as f:
            writer = csv.writer(f)
            writer.writerow(fields)
    def str2bool1(self):
        print("Hello World!")

    def str2bool(self, v):
        opxl=AutoSphereExcel()
        opxl.str2bool1()
        return v.lower() in ("yes", "true", "t", "1")

    def move_cell_range(self, ref, rows, cols):
            # create a table
        ws=self.sheet
#        ref=ref1+":"+ref2
#        isTranslate = (translate.lower() in ("yes", "true", "t", "1"))
        ws.move_range(ref, int(rows), int(cols))    

    def read_csv(self, csvfile, excelfile):
        wb = openpyxl.Workbook()
        ws = wb.active

        with open(csvfile) as f:
            reader = csv.reader(f, delimiter=':')
            for row in reader:
                ws.append(row)

        wb.save(excelfile)
    
    def remove_duplicate_rows(self):
        sh=self.sheet
        rangeSelected = []
        rowindex = []
    
    #thisset.clear()
        for i in range(1,int(sh.max_row) + 1,1):
            print("Processing: " + str(i))
            rowSelected = []
        
            for j in range(1,int(sh.max_column)+1,1):
                rowSelected.append(sh.cell(row = i, column = j).value)
            
            if rowSelected in rangeSelected:
                rowindex.append(i)
#               sh.delete_rows(i,1)          
            else:    
                rangeSelected.append(rowSelected)
        m = 0
        for k in rowindex:
            print(str(k) +" have duplicate entry")
            sh.delete_rows(k-m)
            m = m+1

        return rangeSelected
    def toSingleQuote(s1):
        return "'%s'" % s1                                           
    def splitReference(self, ref):
        split_ref=ref.split(":");
        return (split_ref)

    def convertToRowsColumns(self, ref):
        thistuple =xl_cell_to_rowcol(ref) 
        return thistuple

    def protect_excel_file(self, workbook, password, lock):
        wb = openpyxl.load_workbook(workbook)
        wb.security = WorkbookProtection()
        wb.security.set_workbook_password(password)
        # wb.security.workbookPassword = password
        wb.security.lockStructure = bool(lock)
        wb.save(workbook)

    def protect_sheet(self, workbook, sheetname, password):
        ws=self.sheet
        ws.protection.set_password(password)


    def check_sheet_exists(self, sheetname):
        if sheetname in self.wb.get_sheet_names():
            return True
        return False

    def move_sheet(self, from_loc=None, to_loc=None):
        sheets = self.wb._sheets

        # if no from_loc given, assume last sheet
        if from_loc is None:
            from_loc = len(sheets) - 1

        # if no to_loc given, assume first
        if to_loc is None:
            to_loc = 0

        sheet = sheets.pop(int(from_loc))
        sheets.insert(int(to_loc), sheet)

    def get_cell_color(self, cell):
        i = self.sheet[cell].fill  # Green Color
        print(str(i.fgColor.rgb))
        return i.fgColor.rgb

    def merge_cells(self, cellrange):
        self.sheet.merge_cells(cellrange)

    def unmerge_cells(self, cellrange):
        self.sheet.unmerge_cells(cellrange)

    def format_cell(self, cell, bold_text=False, italic_text=False, text_font='Calibri', text_size=11, text_color='FF000000', text_strike=False):
        # font = Font(bold=bool(bold_text), italic=bool(italic_text), name=text_font, size=int(text_size), color=text_color, strike=bool(text_strike), underline=bool(text_underline))
        font = Font(name=text_font,
                    size=int(text_size),
                    bold=bool(bold_text),
                    italic=bool(italic_text),
                    vertAlign=None,
                    underline='none',
                    strike=bool(text_strike),
                    color=text_color)
        self.sheet[cell].font = font
        # if bold_text is not None:
        #     self.sheet[cell].font = Font(bold=bool(bold_text))
        # if italic_text is not None:
        #     self.sheet[cell].font = Font(italic=bool(italic_text))
        # if text_font is not None:
        #     self.sheet[cell].font = Font(name=text_font)
        # if text_size is not None:
        #     self.sheet[cell].font = Font(size=int(text_size))
        # if text_color is not None:
        #     self.sheet[cell].font = Font(color=text_color)
        # if text_strike is not None:
        #     self.sheet[cell].font = Font(strike=bool(text_strike))
        # if text_underline is not None:
        #     self.sheet[cell].font = Font(underline=bool(text_underline))
        # Font properties
        # _cell.font.color = 'FF00FF00'
        # _cell.font = 'Arial'
        # _cell.font = _cell.font.copy(italic=True)
        # _cell.font.bold = True
        # _cell.style.alignment.wrap_text = True
        #
        # # Cell background color
        # _cell.style.fill.fill_type = Fill.FILL_SOLID
        # _cell.style.fill.start_color.index = Color.DARKRED
        #
        # # You should only modify column dimensions after you have written a cell in
        # #     the column. Perfect world: write column dimensions once per column
        # #
        # ws.column_dimensions["C"].width = 60.0

    def is_read_only(self):
        res = self.wb.read_only
        print("-----------"+str(res))
        return res

    def is_write_only(self):
        res = self.wb.write_only
        print(res)
        return res

    def execute_macro(self, workbook, macro):
        if os.path.exists(workbook):
            xl = win32com.client.Dispatch("Excel.Application")
            wb = xl.workbooks.open(workbook)
            xl.run(macro) # UnhideAllWoksheets, HideAllExceptActiveSheet
            xl.Visible = True
            wb.Close(SaveChanges=1)
            xl.Quit()

#
if __name__ == "__main__":
    opxl=AutoSphereExcel()
    opxl.open_excel('test1.xlsx')
    opxl.get_sheet_names()
    opxl.save_excel('test1.xlsx')
    print(opxl.str2bool('yes'))
    s_ref=(opxl.splitReference("A2:C5"));
    print(s_ref[0]);print(s_ref[1]);
    row_col=(opxl.convertToRowsColumns(s_ref[0]));
    print(row_col[0]);print(row_col[1]);
    row_col=(opxl.convertToRowsColumns(s_ref[1]));
    print(row_col[0]);print(row_col[1]);
    opxl.close()