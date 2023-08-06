#!/usr/bin/python
# -*- coding: UTF-8 -*-

import os
import openpyxl
import xlrd
import xlwt

class Excel:
    _wb = None
    _sheet = None
    _file_name = False
    _keys = []
    _excel_type = 'xls'

    # 初始化
    def __init__(self, file="", hasKey=True, sheet_name=""):
        if file == "":
            return None

        self._file_name = file
        self._excel_type = 'xlsx' if ".xlsx" in self._file_name else self._excel_type

        if os.path.isfile(self._file_name):
            if self._excel_type == 'xlsx':
                self._wb = openpyxl.load_workbook(self._file_name)
                self._sheet = self._wb.active
            else:
                # self._wb = self._open_xls_as_xlsx(self._file_name)
                self._wb = xlrd.open_workbook(self._file_name)
                self._sheet = self._wb.sheets()[0]
        else:
            if self._excel_type == 'xlsx':
                self._wb = openpyxl.Workbook()
                self._sheet = self._wb.active
            else:
                self._wb = xlwt.Workbook()
                if sheet_name == "":
                    sheet_name = 'Sheet1'
                self._sheet = self._wb.add_sheet(sheet_name)

        # self._sheet = self._wb.active if self._sheet == None else self._sheet
        if hasKey:
            self.parse_key()

    # 解析标题栏
    def parse_key(self):
        start = 1 if self._excel_type == 'xlsx' else 0
        for index in range(self.get_cols()):
            self._keys.append(self._sheet.cell(start, index+start).value)

    def get_titles(self):
        return self._keys

    # 根据行/列读写数据
    def column(self, row, col, value=None):
        # col_letter = get_column_letter(col)
        if value == None:
            result = self._sheet.cell(row, col).value
            if type(result) is type(''):
                result = result.replace('_x000D_','')
            return result

        try:
            if self._excel_type == "xlsx":
                self._sheet.cell(row, col).value = value
            else:
                self._sheet.write(row-1, col-1, value)
        except:
            pass

    # 根据行及key读取数据
    def get_value(self, row, key, default=""):
        row = row if self._excel_type == "xlsx" else row -1

        try:
            if key in self._keys:
                col = self._keys.index(key) + 1 if self._excel_type == "xlsx" else self._keys.index(key)
                row += 1

                result = self._sheet.cell(row, col).value
                if type(result) is type(''):
                    result = result.replace('_x000D_','')
                
                if default == 0:
                    return 0 if result == "" else int(result)
                if default == "":
                    return "" if result == None else str(result)

                return result
        except:
            pass
        
        return default
        
    # 设置行列数据
    def set_value(self, row, key, value):
        if key in self._keys:
            col = self._keys.index(key) + 1
            self._sheet.cell(row, col).value = value
            return True

        return False

    # 获取当前工作表的标题
    def title(self):
        return self._sheet.title

    # 获取当前工作表行数
    def get_rows(self):
        if self._excel_type == "xlsx":
            return self._sheet.max_row - 1 if self._sheet.max_row > 0 else 0
        else:
            return self._sheet.nrows - 1 if self._sheet.nrows > 0 else 0

    # 获取当前工作表列数
    def get_cols(self):
        if self._excel_type == 'xlsx':
            return self._sheet.max_column
        else:
            return self._sheet.ncols

    # 存储Excel
    def save(self):
        self._wb.save(self._file_name)

    # 打开xls转化为xlsx-暂时弃用
    def _open_xls_as_xlsx(self, filename):
        # first open using xlrd
        xls_book = xlrd.open_workbook(filename)
        sheet = xls_book.sheets()[0]
        nrows = sheet.nrows
        ncols = sheet.ncols

        # 读取多个sheet，暂时不支持
        # index = 0
        # nrows, ncols = 0, 0
        # while nrows * ncols == 0:
        #     sheet = xls_book.sheet_by_index(index)
        #     nrows = sheet.nrows
        #     ncols = sheet.ncols
        #     index += 1

        # prepare a xlsx sheet
        xlsx_book = openpyxl.Workbook()
        xlsx_sheet = xlsx_book.get_active_sheet()

        for row in range(0, nrows):
            for col in range(0, ncols):
                xlsx_sheet.cell(row+1, col+1).value = sheet.cell_value(row, col)

        return xlsx_book
